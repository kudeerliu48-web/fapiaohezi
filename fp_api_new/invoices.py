import os
import re
import sqlite3
import shutil
import threading
import time
import uuid
import imaplib
import email as pyemail
from email.header import decode_header
from email.utils import parsedate_to_datetime
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple, Union
from pathlib import Path
import io
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import aiofiles
import requests
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import json

from utils import (
    generate_uuid, format_datetime, safe_json_dumps, safe_json_loads,
    get_file_extension, is_allowed_file, ResponseHelper
)
from config import config

try:
    from fp_api_new.user import check_recognition_quota
except Exception:
    from user import check_recognition_quota

router = APIRouter(prefix="/api", tags=["发票管理"])

MAX_UPLOAD_FILE_SIZE = 300 * 1024

_email_jobs: Dict[str, Dict[str, Any]] = {}
_email_jobs_lock = threading.Lock()
_recognition_jobs: Dict[str, Dict[str, Any]] = {}
_recognition_jobs_lock = threading.Lock()


class DeriveInvoiceBody(BaseModel):
    """基于原发票新增一条可编辑记录的请求体（发票号不可改，其余可编辑）"""
    invoice_date: Optional[str] = None
    service_name: Optional[str] = None
    amount_without_tax: Optional[float] = None
    tax_amount: Optional[float] = None
    total_with_tax: Optional[float] = None
    buyer: Optional[str] = None
    seller: Optional[str] = None
    remark: Optional[str] = None  # 备注，存 field1


def get_main_db_connection():
    conn = sqlite3.connect("main.db")
    conn.row_factory = sqlite3.Row
    return conn


def get_user_db_connection(user_id: str):
    user_db_path = config.get_user_db_path(user_id)
    os.makedirs(os.path.dirname(user_db_path), exist_ok=True)
    init_user_database(user_id, user_db_path)
    conn = sqlite3.connect(user_db_path)
    conn.row_factory = sqlite3.Row
    return conn


def _sync_invoice_to_main(user_id: str, invoice_id: str) -> None:
    """将用户库中的一条发票同步到 main.db 的 invoice_sync 表，供管理端查询。"""
    try:
        uconn = get_user_db_connection(user_id)
        uc = uconn.cursor()
        uc.execute("SELECT * FROM invoice_details WHERE id = ?", (invoice_id,))
        row = uc.fetchone()
        uconn.close()
        if not row:
            return
        r = dict(row)
        main_conn = get_main_db_connection()
        mc = main_conn.cursor()
        sync_id = f"{user_id}_{invoice_id}"
        now = format_datetime()
        mc.execute(
            """INSERT OR REPLACE INTO invoice_sync (
                id, user_id, invoice_id, batch_id, invoice_number, invoice_date,
                invoice_amount, total_with_tax, buyer, seller, recognition_status,
                processing_time, upload_time, file_type, file_size, ocr_text,
                filename, saved_filename, updated_at, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                sync_id,
                user_id,
                invoice_id,
                r.get("batch_id"),
                r.get("invoice_number"),
                r.get("invoice_date"),
                r.get("invoice_amount") or r.get("total_with_tax"),
                r.get("total_with_tax"),
                r.get("buyer"),
                r.get("seller"),
                r.get("recognition_status"),
                r.get("processing_time"),
                r.get("upload_time"),
                r.get("file_type"),
                r.get("file_size"),
                r.get("ocr_text"),
                r.get("filename"),
                r.get("saved_filename"),
                r.get("updated_at") or now,
                r.get("created_at") or now,
            ),
        )
        main_conn.commit()
        # 同步更新该用户的已用识别次数：以 invoice_sync 中 recognition_status=1 的数量为准
        mc.execute(
            """UPDATE users SET recognition_used = (
                SELECT COUNT(*) FROM invoice_sync WHERE user_id = ? AND recognition_status = 1
            ) WHERE id = ?""",
            (user_id, user_id),
        )
        main_conn.commit()
        main_conn.close()
    except Exception as e:
        print(f"[sync_invoice_to_main] {user_id} {invoice_id}: {e}")


def init_user_database(user_id: str, db_path: str) -> bool:
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoice_details (
                id TEXT PRIMARY KEY,
                batch_id TEXT,
                filename TEXT NOT NULL,
                saved_filename TEXT,
                processed_filename TEXT,
                color_filename TEXT,
                original_file_path TEXT,
                processed_file_path TEXT,
                page_index INTEGER,
                invoice_amount REAL,
                buyer TEXT,
                seller TEXT,
                invoice_number TEXT,
                invoice_date TEXT,
                service_name TEXT,
                amount_without_tax REAL,
                tax_amount REAL,
                total_with_tax REAL,
                final_json TEXT,
                total_duration_ms REAL,
                recognition_status INTEGER DEFAULT 0,
                processing_time REAL,
                ocr_text TEXT,
                json_info TEXT,
                file_type TEXT,
                file_size INTEGER,
                upload_time TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                field1 TEXT,
                field2 TEXT,
                field3 TEXT,
                source_invoice_id TEXT
            )
        ''')
        try:
            cursor.execute("PRAGMA table_info(invoice_details)")
            cols = [r[1] for r in cursor.fetchall()]
            if 'source_invoice_id' not in cols:
                cursor.execute("ALTER TABLE invoice_details ADD COLUMN source_invoice_id TEXT")
        except Exception:
            pass

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS batches (
                id TEXT PRIMARY KEY,
                user_id TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'processing',
                total_invoices INTEGER DEFAULT 0,
                success_count INTEGER DEFAULT 0,
                failed_count INTEGER DEFAULT 0,
                total_duration_ms REAL DEFAULT 0,
                remark TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoice_steps (
                id TEXT PRIMARY KEY,
                invoice_id TEXT NOT NULL,
                batch_id TEXT,
                step_name TEXT NOT NULL,
                step_order INTEGER DEFAULT 0,
                status TEXT NOT NULL,
                started_at TEXT,
                ended_at TEXT,
                duration_ms REAL,
                input_payload TEXT,
                output_payload TEXT,
                error_message TEXT,
                debug_meta TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        ''')

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_invoice_details_batch_id ON invoice_details(batch_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_invoice_details_upload_time ON invoice_details(upload_time)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_batches_created_at ON batches(created_at)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_invoice_steps_invoice_id ON invoice_steps(invoice_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_invoice_steps_batch_id ON invoice_steps(batch_id)")
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_invoice_steps_invoice_step_unique ON invoice_steps(invoice_id, step_name)")

        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"初始化用户数据库失败: {e}")
        return False


def _generate_batch_id(user_id: str) -> str:
    """
    生成批次号：用户ID + 当天年月日(yyMMdd) + 当前时间戳（秒）。
    示例：<user_id>260313<timestamp>
    """
    date_key = datetime.now().strftime("%y%m%d")  # 如 2026-03-13 -> 260313
    ts = int(datetime.now().timestamp())
    return f"{user_id}{date_key}{ts}"


def _ensure_user_exists(user_id: str) -> bool:
    conn = get_main_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE id = ?", (user_id,))
    exists = cursor.fetchone() is not None
    conn.close()
    return exists


def _resolve_batch_owner_user_id(batch_id: str) -> Optional[str]:
    """
    在 files/{user_id}/database/{user_id}.db 中查 batches.id 反查所属用户。
    如果找不到则返回 None。
    """
    base_dir = config.UPLOAD_DIR
    if not os.path.exists(base_dir):
        return None

    for uid in os.listdir(base_dir):
        user_db = config.get_user_db_path(uid)
        if not os.path.exists(user_db):
            continue
        try:
            conn = sqlite3.connect(user_db)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM batches WHERE id = ? LIMIT 1", (batch_id,))
            row = cursor.fetchone()
            conn.close()
            if row:
                return uid
        except Exception:
            try:
                conn.close()
            except Exception:
                pass
            continue
    return None


def _update_invoice_from_stream_item(owner_user_id: str, item: Dict[str, Any]) -> None:
    """
    将 V9 stream 的 invoice_item 写回用户库 invoice_details。
    - 优先按 batch_id + filename 匹配
    - 其次按 batch_id + saved_filename 匹配
    """
    batch_id = str(item.get("batch_id") or "").strip()
    file_name = str(item.get("file_name") or "").strip()
    if not batch_id or not file_name:
        return

    status = str(item.get("status") or "").strip().lower()
    if status == "success":
        rec_status = 1
    elif status == "failed":
        rec_status = 2
    else:
        rec_status = 0

    invoice_number = item.get("invoice_number")
    invoice_date = item.get("issue_date") or item.get("invoice_date")
    buyer = item.get("buyer_name") or item.get("buyer")
    seller = item.get("seller_name") or item.get("seller")
    service_name = item.get("service_name")
    amount_without_tax = item.get("total_amount") or item.get("amount_without_tax")
    tax_amount = item.get("tax_amount")
    total_with_tax = item.get("amount_with_tax") or item.get("total_with_tax")

    now = format_datetime()
    raw_json = safe_json_dumps(item)

    conn = get_user_db_connection(owner_user_id)
    cursor = conn.cursor()

    cursor.execute(
        """
        UPDATE invoice_details
        SET recognition_status = ?,
            invoice_number = COALESCE(?, invoice_number),
            invoice_date = COALESCE(?, invoice_date),
            buyer = COALESCE(?, buyer),
            seller = COALESCE(?, seller),
            service_name = COALESCE(?, service_name),
            amount_without_tax = COALESCE(?, amount_without_tax),
            tax_amount = COALESCE(?, tax_amount),
            total_with_tax = COALESCE(?, total_with_tax),
            invoice_amount = COALESCE(?, invoice_amount),
            json_info = ?,
            final_json = ?,
            updated_at = ?
        WHERE batch_id = ? AND filename = ?
        """,
        (
            rec_status,
            invoice_number,
            invoice_date,
            buyer,
            seller,
            service_name,
            amount_without_tax,
            tax_amount,
            total_with_tax,
            total_with_tax,
            raw_json,
            raw_json,
            now,
            batch_id,
            file_name,
        ),
    )

    if cursor.rowcount == 0:
        cursor.execute(
            """
            UPDATE invoice_details
            SET recognition_status = ?,
                invoice_number = COALESCE(?, invoice_number),
                invoice_date = COALESCE(?, invoice_date),
                buyer = COALESCE(?, buyer),
                seller = COALESCE(?, seller),
                service_name = COALESCE(?, service_name),
                amount_without_tax = COALESCE(?, amount_without_tax),
                tax_amount = COALESCE(?, tax_amount),
                total_with_tax = COALESCE(?, total_with_tax),
                invoice_amount = COALESCE(?, invoice_amount),
                json_info = ?,
                final_json = ?,
                updated_at = ?
            WHERE batch_id = ? AND saved_filename = ?
            """,
            (
                rec_status,
                invoice_number,
                invoice_date,
                buyer,
                seller,
                service_name,
                amount_without_tax,
                tax_amount,
                total_with_tax,
                total_with_tax,
                raw_json,
                raw_json,
                now,
            batch_id,
            file_name,
            ),
        )

    conn.commit()
    cursor.execute(
        "SELECT id FROM invoice_details WHERE batch_id = ? AND (filename = ? OR saved_filename = ?)",
        (batch_id, file_name, file_name),
    )
    for row in cursor.fetchall():
        _sync_invoice_to_main(owner_user_id, row["id"])
    conn.close()


def _update_batch_from_done(owner_user_id: str, batch_id: str, done_payload: Dict[str, Any]) -> None:
    status = str(done_payload.get("status") or "").strip().lower()
    mapped = "completed" if status == "completed" else ("failed" if status == "failed" else status or "completed")
    now = format_datetime()

    conn = get_user_db_connection(owner_user_id)
    cursor = conn.cursor()

    cursor.execute(
        "SELECT COUNT(*) as total, SUM(CASE WHEN recognition_status = 1 THEN 1 ELSE 0 END) as success, SUM(CASE WHEN recognition_status = 2 THEN 1 ELSE 0 END) as failed FROM invoice_details WHERE batch_id = ?",
        (batch_id,),
    )
    row = cursor.fetchone()
    total = int(row[0] or 0) if row else 0
    success = int(row[1] or 0) if row else 0
    failed = int(row[2] or 0) if row else 0

    cursor.execute(
        "UPDATE batches SET status = ?, total_invoices = ?, success_count = ?, failed_count = ?, updated_at = ? WHERE id = ?",
        (mapped, total, success, failed, now, batch_id),
    )
    conn.commit()
    conn.close()


@router.post("/invoices/upload/{user_id}")
async def upload_files(
    user_id: str,
    files: List[UploadFile] = File(...),
):
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)

    if not files:
        return ResponseHelper.error("请至少上传一个文件", 400)

    valid_files = []
    failed_files = []

    for file in files:
        content = await file.read()
        file_size = len(content)

        if file_size > MAX_UPLOAD_FILE_SIZE:
            failed_files.append({
                "filename": file.filename,
                "reason": f"文件超过300KB限制，当前大小: {file_size / 1024:.1f}KB"
            })
            continue

        ext = get_file_extension(file.filename)
        if not is_allowed_file(file.filename, config.ALLOWED_EXTENSIONS):
            failed_files.append({
                "filename": file.filename,
                "reason": f"不支持的文件类型: {ext}"
            })
            continue

        valid_files.append({
            "filename": file.filename,
            "content": content,
            "file_size": file_size,
            "ext": ext
        })

    if not valid_files:
        return ResponseHelper.error("没有有效的文件可上传", 400)

    batch_id = _generate_batch_id(user_id)
    user_db_path = config.get_user_db_path(user_id)
    upload_dir = config.get_upload_dir(user_id)
    os.makedirs(upload_dir, exist_ok=True)

    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()

    now = format_datetime()
    cursor.execute('''
        INSERT INTO batches (
            id, user_id, status, total_invoices, success_count, failed_count,
            total_duration_ms, remark, created_at, updated_at
        ) VALUES (?, ?, ?, 0, 0, 0, 0, ?, ?, ?)
    ''', (batch_id, user_id, "pending", "", now, now))
    conn.commit()

    created_invoices = []
    for file_info in valid_files:
        file_id = generate_uuid()
        saved_filename = f"{file_id}{file_info['ext']}"
        file_path = os.path.join(upload_dir, saved_filename)

        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(file_info['content'])

        invoice_id = generate_uuid()
        cursor.execute('''
            INSERT INTO invoice_details (
                id, batch_id, filename, saved_filename, original_file_path,
                file_type, file_size, upload_time, recognition_status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            invoice_id, batch_id, file_info['filename'], saved_filename,
            f"/files/{user_id}/uploads/{saved_filename}",
            file_info['ext'], file_info['file_size'], now, 0, now, now
        ))

        created_invoices.append({
            "id": invoice_id,
            "filename": file_info['filename'],
            "saved_filename": saved_filename,
            "file_size": file_info['file_size'],
            "file_type": file_info['ext'],
            "batch_id": batch_id,
            "recognition_status": 0
        })

    cursor.execute(
        "UPDATE batches SET total_invoices = ? WHERE id = ?",
        (len(created_invoices), batch_id)
    )
    conn.commit()
    for inv in created_invoices:
        _sync_invoice_to_main(user_id, inv["id"])
    conn.close()

    return ResponseHelper.success({
        "batch_id": batch_id,
        "user_id": user_id,
        "total_files": len(created_invoices),
        "failed_count": len(failed_files),
        "files": created_invoices,
        "failed_files": failed_files
    }, "文件上传成功")


@router.post("/invoices/submit/{user_id}/{batch_id}")
async def submit_recognition(user_id: str, batch_id: str):
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)

    user_db_path = config.get_user_db_path(user_id)
    if not os.path.exists(user_db_path):
        return ResponseHelper.error("批次不存在", 404)

    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM batches WHERE id = ?", (batch_id,))
    batch = cursor.fetchone()
    if not batch:
        conn.close()
        return ResponseHelper.error("批次不存在", 404)

    cursor.execute(
        "SELECT id, filename, saved_filename, original_file_path FROM invoice_details WHERE batch_id = ?",
        (batch_id,)
    )
    invoices = cursor.fetchall()
    conn.close()

    if not invoices:
        return ResponseHelper.error("批次中没有文件", 400)

    # 额度与会员校验：额度用完或会员到期则禁止识别
    add_count = len(invoices)
    allowed, msg = check_recognition_quota(user_id, add_count=add_count)
    if not allowed:
        return ResponseHelper.error(msg, 403)

    upload_dir = config.get_upload_dir(user_id)
    files_payload = []
    file_paths = []

    for inv in invoices:
        saved_filename = inv["saved_filename"]
        file_path = os.path.join(upload_dir, saved_filename)
        if os.path.exists(file_path):
            file_paths.append(file_path)
            files_payload.append({"file_name": inv["filename"]})

    if not file_paths:
        return ResponseHelper.error("没有可识别的文件", 400)

    api_key = config.V9_API_KEY
    submit_url = f"{config.V9_API_BASE_URL}/api/invoices/submit"

    try:
        files_to_send = []
        for fp in file_paths:
            files_to_send.append(("files", (os.path.basename(fp), open(fp, "rb"), "application/octet-stream")))

        data = {"batch_id": batch_id}
        if api_key:
            data["api_key"] = api_key

        resp = requests.post(submit_url, data=data, files=files_to_send, timeout=300)

        for _, (_, fobj, _) in files_to_send:
            try:
                fobj.close()
            except Exception:
                pass

        if resp.status_code >= 400:
            return ResponseHelper.error(f"提交识别失败: {resp.text}", 500)

        result = resp.json()
        if not result.get("success"):
            return ResponseHelper.error(f"提交识别业务失败: {result.get('message', resp.text)}", 500)

        conn = get_user_db_connection(user_id)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE batches SET status = ? WHERE id = ?",
            ("processing", batch_id)
        )
        conn.commit()
        conn.close()

        return ResponseHelper.success({
            "batch_id": batch_id,
            "submitted_files": len(file_paths),
            "submit_result": result.get("data", result)
        }, "识别任务已提交")

    except Exception as e:
        return ResponseHelper.error(f"提交识别异常: {str(e)}", 500)


@router.get("/invoices/batches/{batch_id}/stream")
async def get_batch_stream(
    batch_id: str,
    user_id: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    keyword: Optional[str] = None,
    file_name: Optional[str] = None,
    recognition_status: Optional[int] = None,
):
    """
    代理转发到 V9 服务的 SSE 流接口
    V9 接口: GET /api/invoices/batches/{batch_id}/stream
    """
    try:
        # 构建 V9 服务 URL
        v9_stream_url = f"{config.V9_API_BASE_URL}/api/invoices/batches/{batch_id}/stream"
        
        # 可选：添加 api_key 到请求头或参数（根据 V9 实际鉴权方式）
        headers = {}
        if config.V9_API_KEY:
            headers["X-API-Key"] = config.V9_API_KEY
        
        # 使用 requests 流式请求 V9
        resp = requests.get(v9_stream_url, headers=headers, stream=True, timeout=300)
        
        if resp.status_code >= 400:
            error_text = resp.text if resp.text else f"V9 服务返回错误: {resp.status_code}"
            return ResponseHelper.error(error_text, resp.status_code)
        
        owner_user_id = (user_id or "").strip() or _resolve_batch_owner_user_id(batch_id) or None

        # 代理转发 SSE 流，同时解析并落库
        def generate():
            current_event: Optional[str] = None
            data_lines: List[str] = []

            def _flush_event():
                nonlocal current_event, data_lines, owner_user_id
                if not current_event:
                    data_lines = []
                    return
                if not data_lines:
                    current_event = None
                    return

                data_text = "\n".join(data_lines).strip()
                data_lines = []

                if not data_text:
                    current_event = None
                    return

                try:
                    payload = json.loads(data_text)
                except Exception:
                    current_event = None
                    return

                if not owner_user_id:
                    owner_user_id = (user_id or "").strip() or _resolve_batch_owner_user_id(batch_id) or None

                if owner_user_id:
                    try:
                        if current_event == "invoice_item":
                            _update_invoice_from_stream_item(owner_user_id, payload)
                        elif current_event == "batch_done":
                            _update_batch_from_done(owner_user_id, batch_id, payload)
                    except Exception:
                        # 落库失败不影响 SSE 转发
                        pass

                current_event = None

            for line_bytes in resp.iter_lines(chunk_size=1024, decode_unicode=False):
                if line_bytes is None:
                    continue

                # 原样转发（iter_lines 不带换行，这里补回）
                yield line_bytes + b"\n"

                # 解析（以 utf-8 容错解码）
                try:
                    line = line_bytes.decode("utf-8", errors="replace")
                except Exception:
                    line = ""

                if line == "":
                    _flush_event()
                    continue

                if line.startswith("event:"):
                    current_event = line[len("event:") :].strip()
                    continue

                if line.startswith("data:"):
                    data_lines.append(line[len("data:") :].lstrip())
                    continue

            # 结束前 flush 一次
            _flush_event()
        
        return StreamingResponse(
            generate(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",  # 禁用 Nginx 缓冲
            }
        )
        
    except Exception as e:
        return ResponseHelper.error(f"代理 V9 流接口失败: {str(e)}", 500)


# 保留旧接口作为兼容或删除


@router.get("/invoices/{user_id}")
async def get_invoices(
    user_id: str,
    page: int = 1,
    limit: int = 20,
    keyword: Optional[str] = None,
    batch_id: Optional[str] = None,
    recognition_status: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    recognize_date_from: Optional[str] = None,
    recognize_date_to: Optional[str] = None,
):
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)

    user_db_path = config.get_user_db_path(user_id)
    if not os.path.exists(user_db_path):
        return ResponseHelper.success({
            "invoices": [],
            "total": 0,
            "page": page,
            "limit": limit,
            "pages": 0
        }, "获取发票列表成功")

    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()

    where_clauses = []
    params: List[Any] = []

    if keyword:
        like = f"%{keyword}%"
        params.extend([like, like, like, like])
        where_clauses.append("(invoice_number LIKE ? OR buyer LIKE ? OR seller LIKE ? OR filename LIKE ?)")

    if batch_id:
        where_clauses.append("batch_id = ?")
        params.append(batch_id)

    if recognition_status is not None:
        where_clauses.append("recognition_status = ?")
        params.append(recognition_status)

    if date_from:
        where_clauses.append("upload_time >= ?")
        params.append(date_from)

    if date_to:
        where_clauses.append("upload_time <= ?")
        params.append(date_to)

    if recognize_date_from:
        where_clauses.append("updated_at >= ?")
        params.append(recognize_date_from)

    if recognize_date_to:
        where_clauses.append("updated_at <= ?")
        params.append(recognize_date_to)

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    offset = (page - 1) * limit
    cursor.execute(
        f'''SELECT id, batch_id, filename, saved_filename, original_file_path, processed_file_path,
               invoice_amount, buyer, seller,
               invoice_number, invoice_date, service_name, amount_without_tax, tax_amount,
               total_with_tax, recognition_status, processing_time, upload_time, file_type, file_size,
               updated_at as recognize_time, source_invoice_id, field1
        FROM invoice_details {where_sql}
        ORDER BY COALESCE(source_invoice_id, id), id
        LIMIT ? OFFSET ?''',
        (*params, limit, offset)
    )
    invoices = [dict(row) for row in cursor.fetchall()]

    cursor.execute(f"SELECT COUNT(*) as total FROM invoice_details {where_sql}", params)
    total = cursor.fetchone()["total"]

    conn.close()

    return ResponseHelper.success({
        "invoices": invoices,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }, "获取发票列表成功")


@router.get("/invoices/{user_id}/export")
async def export_invoices(
    user_id: str,
    recognize_date_from: Optional[str] = None,
    recognize_date_to: Optional[str] = None,
    recognition_status: Optional[int] = None,
):
    """
    导出发票清单为 Excel(xlsx)。
    - 支持按识别日期范围/状态筛选（与列表一致）
    - 导出字段固定为“发票清单展示字段”，不包含：识别日期、状态
    """
    if not _ensure_user_exists(user_id):
        raise HTTPException(status_code=404, detail="用户不存在")

    user_db_path = config.get_user_db_path(user_id)
    if not os.path.exists(user_db_path):
        raise HTTPException(status_code=404, detail="无发票数据")

    where_clauses: List[str] = []
    params: List[Any] = []
    if recognition_status is not None:
        where_clauses.append("o.recognition_status = ?")
        params.append(recognition_status)
    if recognize_date_from:
        where_clauses.append("o.updated_at >= ?")
        params.append(recognize_date_from)
    if recognize_date_to:
        where_clauses.append("o.updated_at <= ?")
        params.append(recognize_date_to)
    # 导出以新增行为准：有新增行则导新增行，否则导原始行
    effective_condition = (
        "(o.source_invoice_id IS NOT NULL "
        "OR NOT EXISTS (SELECT 1 FROM invoice_details d WHERE d.source_invoice_id = o.id))"
    )
    where_clauses.append(effective_condition)
    where_sql = f"WHERE {' AND '.join(where_clauses)}"

    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()
    cursor.execute(
        f"""
        SELECT
            o.invoice_number,
            o.invoice_date,
            o.service_name,
            o.amount_without_tax,
            o.total_with_tax,
            o.buyer,
            o.seller,
            o.field1
        FROM invoice_details o
        {where_sql}
        ORDER BY COALESCE(o.source_invoice_id, o.id), o.id
        """,
        params,
    )
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "发票清单"

    headers_cn = ["发票号码", "开票日期", "服务名称", "不含税额", "价税合计", "购买方", "销售方", "备注"]
    keys = ["invoice_number", "invoice_date", "service_name", "amount_without_tax", "total_with_tax", "buyer", "seller", "field1"]

    ws.append(headers_cn)
    for r in rows:
        ws.append([(r.get(k) or "") for k in keys])

    # 简单自适应列宽
    for idx, header in enumerate(headers_cn, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = max(12, len(header) + 2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    # HTTP 头部要求 latin-1，可读性优先选择 ASCII 文件名，内容内再用中文表头
    filename = f"invoices_{user_id}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/invoices/{user_id}/{invoice_id}/derive")
async def derive_invoice(user_id: str, invoice_id: str, body: DeriveInvoiceBody):
    """基于原发票新增一条记录：原行不变，新记录排在原行下，状态=99（黄色问号），不可再编辑。"""
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)
    user_db_path = config.get_user_db_path(user_id)
    if not os.path.exists(user_db_path):
        return ResponseHelper.error("用户数据库不存在", 404)
    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM invoice_details WHERE id = ?", (invoice_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return ResponseHelper.error("原发票不存在", 404)
    orig = dict(row)
    # 不允许基于“新增行”再派生（新增行 recognition_status=99）
    if orig.get("recognition_status") == 99:
        conn.close()
        return ResponseHelper.error("新增行不允许再修改或派生", 400)
    # 检查是否已有基于该行的新增记录
    cursor.execute("SELECT 1 FROM invoice_details WHERE source_invoice_id = ? LIMIT 1", (invoice_id,))
    if cursor.fetchone():
        conn.close()
        return ResponseHelper.error("该原始行已有新增行，不可再编辑", 400)
    now = format_datetime()
    new_id = generate_uuid()
    cursor.execute(
        """INSERT INTO invoice_details (
            id, batch_id, filename, saved_filename, processed_filename, color_filename,
            original_file_path, processed_file_path, page_index, invoice_amount,
            buyer, seller, invoice_number, invoice_date, service_name,
            amount_without_tax, tax_amount, total_with_tax, final_json, total_duration_ms,
            recognition_status, processing_time, ocr_text, json_info, file_type, file_size,
            upload_time, created_at, updated_at, field1, field2, field3, source_invoice_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            new_id,
            orig.get("batch_id"),
            orig.get("filename") or "",
            None,
            None,
            None,
            None,
            None,
            None,
            body.amount_without_tax if body.amount_without_tax is not None else orig.get("invoice_amount"),
            body.buyer if body.buyer is not None else orig.get("buyer"),
            body.seller if body.seller is not None else orig.get("seller"),
            orig.get("invoice_number") or "",
            body.invoice_date if body.invoice_date is not None else orig.get("invoice_date"),
            body.service_name if body.service_name is not None else orig.get("service_name"),
            body.amount_without_tax if body.amount_without_tax is not None else orig.get("amount_without_tax"),
            body.tax_amount if body.tax_amount is not None else orig.get("tax_amount"),
            body.total_with_tax if body.total_with_tax is not None else orig.get("total_with_tax"),
            None,
            None,
            99,
            None,
            None,
            None,
            None,
            None,
            now,
            now,
            now,
            (body.remark or None),
            None,
            None,
            invoice_id,
        ),
    )
    conn.commit()
    _sync_invoice_to_main(user_id, new_id)
    conn.close()
    return ResponseHelper.success({"id": new_id, "source_invoice_id": invoice_id}, "新增记录成功")


@router.put("/invoices/{user_id}/{invoice_id}")
async def update_invoice(user_id: str, invoice_id: str, body: DeriveInvoiceBody):
    """仅允许编辑新增行（recognition_status=99）：更新可编辑字段及备注 field1。"""
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)
    user_db_path = config.get_user_db_path(user_id)
    if not os.path.exists(user_db_path):
        return ResponseHelper.error("用户数据库不存在", 404)
    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()
    cursor.execute("SELECT id, recognition_status FROM invoice_details WHERE id = ?", (invoice_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return ResponseHelper.error("发票不存在", 404)
    if row["recognition_status"] != 99:
        conn.close()
        return ResponseHelper.error("仅允许编辑新增行", 400)
    # 只更新可编辑字段，发票号不改
    updates = []
    params: List[Any] = []
    if body.invoice_date is not None:
        updates.append("invoice_date = ?")
        params.append(body.invoice_date)
    if body.service_name is not None:
        updates.append("service_name = ?")
        params.append(body.service_name)
    if body.amount_without_tax is not None:
        updates.append("amount_without_tax = ?")
        params.append(body.amount_without_tax)
    if body.tax_amount is not None:
        updates.append("tax_amount = ?")
        params.append(body.tax_amount)
    if body.total_with_tax is not None:
        updates.append("total_with_tax = ?")
        params.append(body.total_with_tax)
    if body.buyer is not None:
        updates.append("buyer = ?")
        params.append(body.buyer)
    if body.seller is not None:
        updates.append("seller = ?")
        params.append(body.seller)
    if body.remark is not None:
        updates.append("field1 = ?")
        params.append(body.remark)
    if not updates:
        conn.close()
        return ResponseHelper.success({"updated": True}, "无变更")
    updates.append("updated_at = ?")
    params.append(format_datetime())
    params.append(invoice_id)
    cursor.execute(
        f"UPDATE invoice_details SET {', '.join(updates)} WHERE id = ?",
        params,
    )
    conn.commit()
    conn.close()
    return ResponseHelper.success({"updated": True}, "更新成功")


class RemarkBody(BaseModel):
    remark: Optional[str] = None


@router.patch("/invoices/{user_id}/{invoice_id}/remark")
async def update_invoice_remark(user_id: str, invoice_id: str, body: RemarkBody):
    """仅更新发票备注（field1），任意行均可."""
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)
    user_db_path = config.get_user_db_path(user_id)
    if not os.path.exists(user_db_path):
        return ResponseHelper.error("用户数据库不存在", 404)
    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM invoice_details WHERE id = ?", (invoice_id,))
    if not cursor.fetchone():
        conn.close()
        return ResponseHelper.error("发票不存在", 404)
    remark_val = body.remark if body.remark is not None else ""
    cursor.execute(
        "UPDATE invoice_details SET field1 = ?, updated_at = ? WHERE id = ?",
        (remark_val, format_datetime(), invoice_id),
    )
    conn.commit()
    conn.close()
    return ResponseHelper.success({"updated": True}, "备注已保存")


@router.delete("/invoices/{user_id}/{invoice_id}")
async def delete_invoice(user_id: str, invoice_id: str):
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)

    user_db_path = config.get_user_db_path(user_id)
    if not os.path.exists(user_db_path):
        return ResponseHelper.error("发票不存在", 404)

    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM invoice_details WHERE id = ?", (invoice_id,))
    invoice = cursor.fetchone()
    if not invoice:
        conn.close()
        return ResponseHelper.error("发票不存在", 404)

    invoice_dict = dict(invoice)

    upload_dir = config.get_upload_dir(user_id)
    if invoice_dict.get("saved_filename"):
        file_path = os.path.join(upload_dir, invoice_dict["saved_filename"])
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception:
                pass

    cursor.execute("DELETE FROM invoice_details WHERE id = ?", (invoice_id,))
    cursor.execute("DELETE FROM invoice_steps WHERE invoice_id = ?", (invoice_id,))

    if invoice_dict.get("batch_id"):
        cursor.execute(
            "SELECT COUNT(*) as total, SUM(CASE WHEN recognition_status = 1 THEN 1 ELSE 0 END) as success, SUM(CASE WHEN recognition_status = 2 THEN 1 ELSE 0 END) as failed FROM invoice_details WHERE batch_id = ?",
            (invoice_dict["batch_id"],)
        )
        stats = cursor.fetchone()
        cursor.execute(
            "UPDATE batches SET total_invoices = ?, success_count = ?, failed_count = ?, updated_at = ? WHERE id = ?",
            (stats["total"] or 0, stats["success"] or 0, stats["failed"] or 0, format_datetime(), invoice_dict["batch_id"])
        )

    conn.commit()
    conn.close()
    try:
        main_conn = get_main_db_connection()
        mc = main_conn.cursor()
        mc.execute("DELETE FROM invoice_sync WHERE id = ?", (f"{user_id}_{invoice_id}",))
        mc.execute(
            """UPDATE users SET recognition_used = (
                SELECT COUNT(*) FROM invoice_sync WHERE user_id = ? AND recognition_status = 1
            ) WHERE id = ?""",
            (user_id, user_id),
        )
        main_conn.commit()
        main_conn.close()
    except Exception:
        pass
    return ResponseHelper.success({"deleted": True}, "发票删除成功")


@router.delete("/invoices/batch/{user_id}/{batch_id}")
async def delete_batch(user_id: str, batch_id: str):
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)

    user_db_path = config.get_user_db_path(user_id)
    if not os.path.exists(user_db_path):
        return ResponseHelper.error("批次不存在", 404)

    conn = get_user_db_connection(user_id)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM batches WHERE id = ?", (batch_id,))
    batch = cursor.fetchone()
    if not batch:
        conn.close()
        return ResponseHelper.error("批次不存在", 404)

    cursor.execute("SELECT saved_filename FROM invoice_details WHERE batch_id = ?", (batch_id,))
    invoices = cursor.fetchall()

    upload_dir = config.get_upload_dir(user_id)
    for inv in invoices:
        if inv["saved_filename"]:
            file_path = os.path.join(upload_dir, inv["saved_filename"])
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass

    cursor.execute("DELETE FROM invoice_details WHERE batch_id = ?", (batch_id,))
    cursor.execute("DELETE FROM invoice_steps WHERE batch_id = ?", (batch_id,))
    cursor.execute("DELETE FROM batches WHERE id = ?", (batch_id,))

    conn.commit()
    conn.close()

    return ResponseHelper.success({"deleted": True}, "批次删除成功")


_EMAIL_TERMINAL_STATUS = {"completed", "failed", "partial_success", "cancelled"}
_EMAIL_STAGE_BASE_PROGRESS = {
    "queued": 0.0,
    "connect_imap": 8.0,
    "select_mailbox": 15.0,
    "search_emails": 25.0,
    "filter_emails": 35.0,
    "parse_message": 48.0,
    "download_attachment": 62.0,
    "import_invoice": 75.0,
    "finalize": 96.0,
}


def _email_now() -> str:
    return datetime.now().isoformat()


def _safe_filename(name: str) -> str:
    name = name.strip().replace("\\", "_").replace("/", "_")
    name = re.sub(r"\s+", " ", name)
    return name or f"file_{uuid.uuid4().hex}"


def _safe_decode_header(raw: Union[str, bytes]) -> str:
    if not raw:
        return ""
    if isinstance(raw, bytes):
        try:
            raw = raw.decode("utf-8", errors="ignore")
        except Exception:
            raw = str(raw)
    try:
        parts = decode_header(raw)
        out: List[str] = []
        for content, enc in parts:
            if isinstance(content, bytes):
                try:
                    out.append(content.decode(enc or "utf-8", errors="ignore"))
                except Exception:
                    out.append(content.decode("utf-8", errors="ignore"))
            else:
                out.append(str(content))
        return "".join(out)
    except Exception:
        return str(raw)


def _imap_host_for_email(addr: str) -> str:
    domain = (addr.split("@", 1)[-1] or "").lower()
    if domain in {"qq.com", "foxmail.com"}:
        return "imap.qq.com"
    if domain == "163.com":
        return "imap.163.com"
    if domain == "126.com":
        return "imap.126.com"
    if domain == "gmail.com":
        return "imap.gmail.com"
    return f"imap.{domain}" if domain else "imap.qq.com"


def _normalize_email_job_payload(job_id: Optional[str], job: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not job_id or not job:
        return {
            "job_id": job_id,
            "task_type": "email_pull",
            "status": "not_found",
            "batch_id": None,
            "user_id": None,
            "mailbox_account": "",
            "days": 0,
            "current_stage": "queued",
            "scanned_emails": 0,
            "matched_emails": 0,
            "downloaded_attachments": 0,
            "imported_invoices": 0,
            "failed_count": 0,
            "progress_percent": 0.0,
            "started_at": None,
            "finished_at": None,
            "logs": [],
            "errors": [],
        }

    raw_status = str(job.get("status") or "queued")
    status = raw_status
    if raw_status == "error":
        status = "failed"

    progress = float(job.get("progress_percent") or 0.0)
    if status in _EMAIL_TERMINAL_STATUS:
        progress = 100.0

    return {
        "job_id": job_id,
        "task_type": "email_pull",
        "status": status,
        "raw_status": raw_status,
        "batch_id": job.get("batch_id"),
        "user_id": job.get("user_id"),
        "mailbox_account": job.get("mailbox_account") or "",
        "days": int(job.get("days") or 0),
        "current_stage": job.get("current_stage") or "queued",
        "scanned_emails": int(job.get("scanned_emails") or 0),
        "matched_emails": int(job.get("matched_emails") or 0),
        "downloaded_attachments": int(job.get("downloaded_attachments") or 0),
        "imported_invoices": int(job.get("imported_invoices") or 0),
        "failed_count": int(job.get("failed_count") or 0),
        "progress_percent": round(progress, 2),
        "started_at": job.get("started_at"),
        "finished_at": job.get("finished_at"),
        "logs": list(job.get("logs") or []),
        "errors": list(job.get("errors") or []),
    }


# def _run_email_pull_job(
#     job_id: str,
#     user_id: str,
#     mailbox: str,
#     auth_code: str,
#     days: int,
#     start_date: Optional[str] = None,
#     end_date: Optional[str] = None,
# ):
#     def _set(**kwargs):
#         with _email_jobs_lock:
#             job = _email_jobs.get(job_id)
#             if job:
#                 job.update(kwargs)
#                 job["updated_at"] = _email_now()

#     def _inc(**delta):
#         with _email_jobs_lock:
#             job = _email_jobs.get(job_id)
#             if job:
#                 for key, value in delta.items():
#                     job[key] = int(job.get(key) or 0) + int(value)

#     def _log(message: str):
#         with _email_jobs_lock:
#             job = _email_jobs.get(job_id)
#             if job:
#                 logs: List[str] = list(job.get("logs") or [])
#                 logs.append(f"{datetime.now().strftime('%H:%M:%S')} {message}")
#                 job["logs"] = logs[-300:]

#     def _error(message: str):
#         with _email_jobs_lock:
#             job = _email_jobs.get(job_id)
#             if job:
#                 errors: List[str] = list(job.get("errors") or [])
#                 errors.append(f"{datetime.now().strftime('%H:%M:%S')} {message}")
#                 job["errors"] = errors[-100:]

#     try:
#         _set(status="running", current_stage="connect_imap")

#         imap_host = _imap_host_for_email(mailbox)
#         imap = imaplib.IMAP4_SSL(imap_host)
#         imap.login(mailbox, auth_code)

#         _set(current_stage="select_mailbox")
#         imap.select("INBOX")

#         today = datetime.now().date()
#         if start_date and end_date:
#             start_day = datetime.strptime(start_date, "%Y-%m-%d").date()
#             end_day = datetime.strptime(end_date, "%Y-%m-%d").date()
#         else:
#             start_day = today - timedelta(days=days)
#             end_day = today

#         _set(current_stage="search_emails")
#         since_date = start_day.strftime("%d-%b-%Y")
#         before_date = (end_day + timedelta(days=1)).strftime("%d-%b-%Y")
#         # 说明：
#         # - imaplib 的 SEARCH 条件默认按 ASCII 编码发送；直接写中文会触发 UnicodeEncodeError
#         # - 这里先按日期范围拉取，再在 Python 中按 Subject 过滤“发票”
#         try:
#             status, data = imap.search(None, f'(SINCE "{since_date}" BEFORE "{before_date}")')
#         except UnicodeEncodeError as e:
#             _error(str(e))
#             status, data = "OK", [b""]

#         if status != "OK":
#             _set(status="failed", finished_at=_email_now())
#             return

#         email_ids = data[0].split()
#         _set(scanned_emails=len(email_ids))

#         _set(current_stage="filter_emails")
#         matched_count = 0

#         upload_dir = config.get_upload_dir(user_id)
#         os.makedirs(upload_dir, exist_ok=True)

#         user_db_path = config.get_user_db_path(user_id)
#         batch_id = _generate_batch_id(user_id)

#         conn = get_user_db_connection(user_id)
#         cursor = conn.cursor()
#         now = format_datetime()
#         cursor.execute('''
#             INSERT INTO batches (
#                 id, user_id, status, total_invoices, success_count, failed_count,
#                 total_duration_ms, remark, created_at, updated_at
#             ) VALUES (?, ?, ?, 0, 0, 0, 0, ?, ?, ?)
#         ''', (batch_id, user_id, "pending", "邮箱拉取", now, now))
#         conn.commit()

#         _set(batch_id=batch_id)

#         for email_id in email_ids:
#             _set(current_stage="parse_message")
#             try:
#                 status, msg_data = imap.fetch(email_id, "(RFC822)")
#                 if status != "OK":
#                     continue
#             except Exception as e:
#                 _error(str(e))
#                 _inc(failed_count=1)
#                 continue

#             msg = pyemail.message_from_bytes(msg_data[0][1])
#             subject = _safe_decode_header(msg.get("Subject", ""))

#             if "发票" not in subject:
#                 continue

#             matched_count += 1
#             _inc(matched_emails=1)

#             _set(current_stage="download_attachment", current_email_subject=subject)

#             for part in msg.walk():
#                 filename = part.get_filename()
#                 if not filename:
#                     continue

#                 decoded_name = _safe_decode_header(filename)
#                 content_type = (part.get_content_type() or "").lower()
#                 ext = os.path.splitext(decoded_name)[1].lower()

#                 if ext not in {".pdf"} and "pdf" not in content_type:
#                     continue

#                 payload = part.get_payload(decode=True)
#                 if not payload:
#                     continue

#                 if len(payload) > MAX_UPLOAD_FILE_SIZE:
#                     _log(f"跳过超大附件: {decoded_name}")
#                     continue

#                 _inc(downloaded_attachments=1)

#                 _set(current_stage="import_invoice", current_attachment_name=decoded_name)

#                 local_name = _safe_filename(f"{uuid.uuid4().hex[:8]}_{decoded_name}")
#                 local_path = os.path.join(upload_dir, local_name)

#                 with open(local_path, "wb") as f:
#                     f.write(payload)

#                 invoice_id = generate_uuid()
#                 cursor.execute('''
#                     INSERT INTO invoice_details (
#                         id, batch_id, filename, saved_filename, original_file_path,
#                         file_type, file_size, upload_time, recognition_status, created_at, updated_at
#                     ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
#                 ''', (
#                     invoice_id, batch_id, decoded_name, local_name,
#                     f"/files/{user_id}/uploads/{local_name}",
#                     ext, len(payload), now, 0, now, now
#                 ))

#                 _inc(imported_invoices=1)
#                 _log(f"导入成功: {decoded_name}")

#         cursor.execute(
#             "UPDATE batches SET total_invoices = ? WHERE id = ?",
#             (int(_email_jobs.get(job_id, {}).get("imported_invoices", 0)), batch_id)
#         )
#         conn.commit()
#         conn.close()

#         imap.logout()

#         _set(status="completed", current_stage="finalize", finished_at=_email_now())

#     except Exception as e:
#         _set(status="failed", finished_at=_email_now())
#         _error(str(e))


# @router.post("/email-push/{user_id}/start")
# async def start_email_push(
#     user_id: str,
#     range_key: str = Form("3m"),
#     mailbox: Optional[str] = Form(None),
#     auth_code: str = Form(...),
#     start_date: Optional[str] = Form(None),
#     end_date: Optional[str] = Form(None),
# ):
#     if not _ensure_user_exists(user_id):
#         return ResponseHelper.error("用户不存在", 404)

#     resolved_mailbox = (mailbox or "").strip()
#     if not resolved_mailbox:
#         conn = get_main_db_connection()
#         cursor = conn.cursor()
#         cursor.execute("SELECT email FROM users WHERE id = ?", (user_id,))
#         user = cursor.fetchone()
#         conn.close()
#         resolved_mailbox = (user["email"] if user else "") or ""

#     if not resolved_mailbox:
#         return ResponseHelper.error("邮箱不能为空", 400)

#     if not auth_code:
#         return ResponseHelper.error("授权码不能为空", 400)

#     range_days_map = {
#         "7d": 7,
#         "1m": 30,
#         "3m": 90,
#         "6m": 180,
#         "12m": 365,
#     }
#     days = range_days_map.get(range_key, 90)

#     job_id = uuid.uuid4().hex
#     with _email_jobs_lock:
#         _email_jobs[job_id] = {
#             "job_id": job_id,
#             "task_type": "email_pull",
#             "status": "queued",
#             "batch_id": None,
#             "user_id": user_id,
#             "mailbox_account": resolved_mailbox,
#             "days": days,
#             "start_date": start_date,
#             "end_date": end_date,
#             "current_stage": "queued",
#             "scanned_emails": 0,
#             "matched_emails": 0,
#             "downloaded_attachments": 0,
#             "imported_invoices": 0,
#             "failed_count": 0,
#             "logs": [],
#             "errors": [],
#             "started_at": _email_now(),
#             "finished_at": None,
#             "progress_percent": 0.0,
#         }

#     t = threading.Thread(
#         target=_run_email_pull_job,
#         args=(job_id, user_id, resolved_mailbox, auth_code, days, start_date, end_date),
#         daemon=True,
#     )
#     t.start()

#     return ResponseHelper.success(_normalize_email_job_payload(job_id, _email_jobs[job_id]), "邮箱拉取任务已开始")





def _run_email_pull_job(
    job_id: str,
    user_id: str,
    mailbox: str,
    auth_code: str,
    days: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    def _set(**kwargs):
        with _email_jobs_lock:
            job = _email_jobs.get(job_id)
            if job:
                job.update(kwargs)
                job["updated_at"] = _email_now()

    def _inc(**delta):
        with _email_jobs_lock:
            job = _email_jobs.get(job_id)
            if job:
                for key, value in delta.items():
                    job[key] = int(job.get(key) or 0) + int(value)

    def _log(message: str):
        with _email_jobs_lock:
            job = _email_jobs.get(job_id)
            if job:
                logs: List[str] = list(job.get("logs") or [])
                logs.append(f"{datetime.now().strftime('%H:%M:%S')} {message}")
                job["logs"] = logs[-300:]

    def _error(message: str):
        with _email_jobs_lock:
            job = _email_jobs.get(job_id)
            if job:
                errors: List[str] = list(job.get("errors") or [])
                errors.append(f"{datetime.now().strftime('%H:%M:%S')} {message}")
                job["errors"] = errors[-100:]

    def _safe_logout(imap_conn):
        try:
            if imap_conn:
                imap_conn.logout()
        except Exception:
            pass

    def _parse_date_range():
        now = datetime.now()

        if start_date:
            try:
                parsed_start = datetime.strptime(start_date, "%Y-%m-%d")
            except ValueError:
                raise ValueError(f"start_date 格式错误: {start_date}")
        else:
            parsed_start = now - timedelta(days=days)

        if end_date:
            try:
                parsed_end = datetime.strptime(end_date, "%Y-%m-%d")
            except ValueError:
                raise ValueError(f"end_date 格式错误: {end_date}")
            parsed_end = parsed_end.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            parsed_end = now

        if parsed_end < parsed_start:
            raise ValueError("结束日期不能早于开始日期")

        # IMAP BEFORE 为开区间，所以要 +1 天
        parsed_before = (parsed_end + timedelta(days=1)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )

        return parsed_start, parsed_end, parsed_before

    def _parse_mail_datetime(msg, internal_date: Optional[str] = None) -> Optional[datetime]:
        """
        优先解析邮件 Date 头；
        如果失败，则回退尝试 INTERNALDATE。
        最终返回可比较的 naive datetime。
        """
        raw_date = msg.get("Date")
        if raw_date:
            try:
                dt = parsedate_to_datetime(raw_date)
                if dt is not None:
                    if dt.tzinfo is not None:
                        dt = dt.astimezone().replace(tzinfo=None)
                    return dt
            except Exception:
                pass

        if internal_date:
            try:
                dt = parsedate_to_datetime(internal_date)
                if dt is not None:
                    if dt.tzinfo is not None:
                        dt = dt.astimezone().replace(tzinfo=None)
                    return dt
            except Exception:
                pass

        return None

    def _extract_internal_date(fetch_response) -> Optional[str]:
        """
        从 FETCH (RFC822 INTERNALDATE) 的响应头里提取 INTERNALDATE 字符串。
        """
        if not fetch_response or not fetch_response[0]:
            return None

        meta = fetch_response[0][0]
        if isinstance(meta, bytes):
            meta = meta.decode(errors="ignore")

        # 示例：
        # b'1 (INTERNALDATE "11-Mar-2026 16:20:31 +0800" RFC822 {1234}'
        marker = 'INTERNALDATE "'
        start = meta.find(marker)
        if start == -1:
            return None
        start += len(marker)
        end = meta.find('"', start)
        if end == -1:
            return None
        return meta[start:end]

    try:
        _set(status="running", current_stage="connect_imap")

        # 1. 建立 IMAP 连接
        imap_host = _imap_host_for_email(mailbox)
        imap = imaplib.IMAP4_SSL(imap_host)
        imap.login(mailbox, auth_code)

        # 2. 只选择收件箱
        _set(current_stage="select_mailbox")
        select_status, _ = imap.select("INBOX")
        if select_status != "OK":
            _set(status="failed", finished_at=_email_now())
            _error("无法选择收件箱 INBOX")
            _safe_logout(imap)
            return

        # 3. 计算日期范围
        try:
            start_day, end_day, before_day = _parse_date_range()
        except ValueError as e:
            _set(status="failed", finished_at=_email_now())
            _error(str(e))
            _safe_logout(imap)
            return

        since_date = start_day.strftime("%d-%b-%Y")
        before_date = before_day.strftime("%d-%b-%Y")

        _set(
            current_stage="search_emails",
            mailbox_name="INBOX",
            range_start=start_day.strftime("%Y-%m-%d %H:%M:%S"),
            range_end=end_day.strftime("%Y-%m-%d %H:%M:%S"),
            imap_search_args=["SINCE", since_date, "BEFORE", before_date],
        )

        _log("当前邮箱文件夹: INBOX")
        _log(
            f"日期范围: {start_day.strftime('%Y-%m-%d %H:%M:%S')} ~ "
            f"{end_day.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        _log(f"IMAP 搜索参数: SINCE {since_date} BEFORE {before_date}")

        # 4. 先取 ALL 做对比，方便判断服务端是否忽略日期条件
        try:
            all_status, all_data = imap.search(None, "ALL")
            inbox_total = len(all_data[0].split()) if all_status == "OK" and all_data and all_data[0] else 0
            _log(f"INBOX 全部邮件数量: {inbox_total}")
        except Exception:
            inbox_total = 0

        # 5. 日期范围搜索（注意：拆参数传递，不拼成整串）
        try:
            status, data = imap.search(None, "SINCE", since_date, "BEFORE", before_date)
        except Exception as e:
            _set(status="failed", finished_at=_email_now())
            _error(f"IMAP 搜索失败: {str(e)}")
            _safe_logout(imap)
            return

        if status != "OK":
            _set(status="failed", finished_at=_email_now())
            _error(f"IMAP 搜索返回失败状态: {status}")
            _safe_logout(imap)
            return

        server_email_ids = data[0].split() if data and data[0] else []
        _log(f"服务器返回 {len(server_email_ids)} 封日期范围候选邮件")

        # 6. 如果服务端搜索明显异常，也不中断，继续靠本地二次过滤兜底
        if not server_email_ids:
            _set(
                scanned_emails=0,
                matched_emails=0,
                downloaded_attachments=0,
                imported_invoices=0,
                status="completed",
                current_stage="finalize",
                finished_at=_email_now(),
                progress_percent=100.0,
            )
            _log("收件箱在该日期范围内没有邮件")
            _safe_logout(imap)
            return

        # 7. 准备目录和数据库（保持你的原逻辑）
        upload_dir = config.get_upload_dir(user_id)
        os.makedirs(upload_dir, exist_ok=True)

        batch_id = _generate_batch_id(user_id)

        conn = get_user_db_connection(user_id)
        cursor = conn.cursor()
        now_ts = format_datetime()

        cursor.execute(
            '''
            INSERT INTO batches (
                id, user_id, status, total_invoices, success_count, failed_count,
                total_duration_ms, remark, created_at, updated_at
            ) VALUES (?, ?, ?, 0, 0, 0, 0, ?, ?, ?)
            ''',
            (
                batch_id,
                user_id,
                "pending",
                f"邮箱拉取({start_day.strftime('%Y-%m-%d')}~{end_day.strftime('%Y-%m-%d')})",
                now_ts,
                now_ts,
            )
        )
        conn.commit()
        _set(batch_id=batch_id)

        # 8. 遍历处理（从新到旧）
        scanned_in_range = 0

        for e_id in reversed(server_email_ids):
            try:
                _set(current_stage="parse_message")

                # 拉完整邮件 + INTERNALDATE
                status, msg_data = imap.fetch(e_id, "(RFC822 INTERNALDATE)")
                if status != "OK" or not msg_data or not msg_data[0]:
                    continue

                raw_bytes = msg_data[0][1]
                if not raw_bytes:
                    continue

                msg = pyemail.message_from_bytes(raw_bytes)
                internal_date = _extract_internal_date(msg_data)

                # 本地二次日期过滤
                mail_dt = _parse_mail_datetime(msg, internal_date=internal_date)
                if mail_dt is None:
                    _error(f"邮件 ID {e_id.decode(errors='ignore')} 缺少或无法解析 Date/INTERNALDATE，已跳过")
                    continue

                if not (start_day <= mail_dt <= end_day):
                    continue

                scanned_in_range += 1
                _set(scanned_emails=scanned_in_range)

                subject = _safe_decode_header(msg.get("Subject", ""))

                # 只处理标题包含“发票”的邮件
                if "发票" not in subject:
                    continue

                _inc(matched_emails=1)
                _log(f"匹配到发票邮件: {subject[:30]}")

                _set(
                    current_stage="download_attachment",
                    current_email_subject=subject,
                )

                # 只处理 PDF 附件
                for part in msg.walk():
                    if part.get_content_maintype() == "multipart":
                        continue

                    filename = part.get_filename()
                    if not filename:
                        continue

                    decoded_name = _safe_decode_header(filename)
                    ext = os.path.splitext(decoded_name)[1].lower()

                    if ext != ".pdf":
                        continue

                    payload = part.get_payload(decode=True)
                    if not payload:
                        continue

                    if len(payload) > MAX_UPLOAD_FILE_SIZE:
                        _error(f"附件过大已跳过: {decoded_name}")
                        continue

                    _inc(downloaded_attachments=1)

                    local_name = _safe_filename(f"{uuid.uuid4().hex[:8]}_{decoded_name}")
                    local_path = os.path.join(upload_dir, local_name)

                    with open(local_path, "wb") as f:
                        f.write(payload)

                    invoice_id = generate_uuid()

                    # 保持你的数据库逻辑不变
                    cursor.execute(
                        '''
                        INSERT INTO invoice_details (
                            id, batch_id, filename, saved_filename, original_file_path,
                            file_type, file_size, upload_time, recognition_status, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''',
                        (
                            invoice_id,
                            batch_id,
                            decoded_name,
                            local_name,
                            f"/files/{user_id}/uploads/{local_name}",
                            ext,
                            len(payload),
                            now_ts,
                            0,
                            now_ts,
                            now_ts,
                        )
                    )
                    conn.commit()

                    _inc(imported_invoices=1)
                    _log(f"导入成功: {decoded_name}")

            except Exception as e:
                _error(f"处理邮件 ID {e_id.decode(errors='ignore')} 出错: {str(e)}")
                _inc(failed_count=1)

        # 9. 收尾
        with _email_jobs_lock:
            final_job = _email_jobs.get(job_id, {}) or {}
            final_count = int(final_job.get("imported_invoices", 0))

        cursor.execute(
            "UPDATE batches SET total_invoices = ?, updated_at = ? WHERE id = ?",
            (final_count, format_datetime(), batch_id)
        )
        conn.commit()
        conn.close()

        _safe_logout(imap)

        _log(f"INBOX 日期范围内共筛到 {scanned_in_range} 封邮件")
        _log(f"最终导入发票 {final_count} 条")

        _set(
            status="completed",
            current_stage="finalize",
            finished_at=_email_now(),
            progress_percent=100.0,
        )

    except Exception as e:
        _set(status="failed", finished_at=_email_now())
        _error(f"严重错误: {str(e)}")

@router.post("/email-push/{user_id}/start")
async def start_email_push(
    user_id: str,
    range_key: str = Form("3m"),
    mailbox: Optional[str] = Form(None),
    auth_code: str = Form(...),
    start_date: Optional[str] = Form(None),
    end_date: Optional[str] = Form(None),
):
    if not _ensure_user_exists(user_id):
        return ResponseHelper.error("用户不存在", 404)

    # 1. 自动获取邮箱
    resolved_mailbox = (mailbox or "").strip()
    if not resolved_mailbox:
        conn = get_main_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT email FROM users WHERE id = ?", (user_id,))
        user = cursor.fetchone()
        conn.close()
        resolved_mailbox = (user["email"] if user else "") or ""

    if not resolved_mailbox:
        return ResponseHelper.error("邮箱不能为空", 400)

    # 2. 映射时间范围
    range_days_map = {
        "7d": 7,
        "1m": 30,
        "3m": 90,
        "6m": 180,
        "12m": 365,
    }
    days = range_days_map.get(range_key, 90)

    # 3. 初始化任务状态
    job_id = uuid.uuid4().hex
    with _email_jobs_lock:
        _email_jobs[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "user_id": user_id,
            "mailbox_account": resolved_mailbox,
            "mailbox_name": "INBOX",
            "days": days,
            "start_date": start_date,
            "end_date": end_date,
            "scanned_emails": 0,          # 日期范围内收件箱邮件数（本地二次过滤后）
            "matched_emails": 0,          # 标题包含发票的邮件数
            "downloaded_attachments": 0,  # 下载的 PDF 附件数
            "imported_invoices": 0,       # 成功导入的发票数
            "failed_count": 0,
            "progress_percent": 0.0,
            "logs": [f"任务已创建，准备拉取 INBOX 最近 {days} 天邮件"],
            "errors": [],
            "started_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat(),
        }

    # 4. 启动后台线程
    t = threading.Thread(
        target=_run_email_pull_job,
        args=(job_id, user_id, resolved_mailbox, auth_code, days, start_date, end_date),
        daemon=True,
    )
    t.start()

    return ResponseHelper.success({"job_id": job_id}, "邮箱拉取任务已开始")


@router.get("/email-push/status/{job_id}")
async def get_email_push_status(job_id: str, user_id: Optional[str] = None):
    with _email_jobs_lock:
        job = _email_jobs.get(job_id)

    return ResponseHelper.success(
        _normalize_email_job_payload(job_id, job),
        "获取邮箱任务状态成功"
    )


@router.get("/email-push/latest/{user_id}")
async def get_latest_email_push(user_id: str):
    with _email_jobs_lock:
        matched = [
            (jid, job)
            for jid, job in _email_jobs.items()
            if job.get("user_id") == user_id
        ]

    if not matched:
        return ResponseHelper.success(
            _normalize_email_job_payload(None, None),
            "无邮箱拉取任务"
        )

    matched.sort(
        key=lambda item: item[1].get("updated_at") or "",
        reverse=True
    )
    job_id, job = matched[0]

    return ResponseHelper.success(
        _normalize_email_job_payload(job_id, job),
        "获取最近邮箱任务成功"
    )
