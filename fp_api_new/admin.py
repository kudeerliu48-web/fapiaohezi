"""
管理端接口：统一使用 main.db，用户表含会员信息（member_type, member_start_at, member_end_at, recognition_used）。
发票数据从 main.db 的 invoice_sync 表读取（用户每处理一张发票会同步到该表）。
"""
import io
import os
import shutil
from datetime import datetime, timedelta
from typing import Any, List, Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils import ResponseHelper, format_datetime
from config import config

try:
    from fp_api_new.user import get_main_db_connection
except Exception:
    from user import get_main_db_connection

router = APIRouter(prefix="/api", tags=["管理端"])


class AdminUserUpdate(BaseModel):
    """管理端更新用户（含会员信息）"""
    username: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    company: Optional[str] = None
    status: Optional[int] = None
    member_type: Optional[str] = None   # normal | free | monthly | yearly
    member_start_at: Optional[str] = None  # 会员开通时间，ISO 格式
    member_end_at: Optional[str] = None     # 会员到期时间（免费用户为开通后3个月）
    recognition_used: Optional[int] = None


@router.get("/admin/users")
async def admin_get_users(page: int = 1, limit: int = 10):
    """获取用户列表（含会员信息），分页。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    offset = (page - 1) * limit
    cursor.execute(
        """SELECT id, username, email, company, phone, status,
                  register_time, last_login_time, avatar_url, role,
                  member_type, member_start_at, member_end_at, recognition_used, created_at, updated_at
           FROM users
           ORDER BY register_time DESC
           LIMIT ? OFFSET ?""",
        (limit, offset),
    )
    rows = cursor.fetchall()
    cursor.execute("SELECT COUNT(*) FROM users")
    total = cursor.fetchone()[0]
    conn.close()
    users = [dict(r) for r in rows]
    for u in users:
        u.setdefault("member_type", "normal")
        u.setdefault("member_start_at", None)
        u.setdefault("member_end_at", None)
        u.setdefault("recognition_used", 0)
    return ResponseHelper.success(
        {
            "users": users,
            "list": users,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit else 0,
        },
        "获取成功",
    )


@router.get("/admin/user/{user_id}")
async def admin_get_user(user_id: str):
    """获取单个用户详情（含会员信息）。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """SELECT id, username, email, company, phone, status,
                  register_time, last_login_time, avatar_url, role,
                  member_type, member_start_at, member_end_at, recognition_used, created_at, updated_at
           FROM users WHERE id = ?""",
        (user_id,),
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return ResponseHelper.error("用户不存在", 404)
    user = dict(row)
    user.setdefault("member_type", "normal")
    user.setdefault("member_start_at", None)
    user.setdefault("member_end_at", None)
    user.setdefault("recognition_used", 0)
    return ResponseHelper.success(user, "获取成功")


@router.put("/admin/user/{user_id}")
async def admin_update_user(user_id: str, body: AdminUserUpdate):
    """更新用户（含会员信息）。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE id = ?", (user_id,))
    if not cursor.fetchone():
        conn.close()
        return ResponseHelper.error("用户不存在", 404)

    updates: List[str] = []
    params: List[Any] = []

    if body.username is not None:
        updates.append("username = ?")
        params.append(body.username.strip())
    if body.phone is not None:
        updates.append("phone = ?")
        params.append(body.phone.strip())
    if body.email is not None:
        updates.append("email = ?")
        params.append(body.email.strip() if body.email else None)
    if body.company is not None:
        updates.append("company = ?")
        params.append(body.company)
    if body.status is not None:
        updates.append("status = ?")
        params.append(body.status)
    if body.member_type is not None:
        if body.member_type not in ("normal", "free", "monthly", "yearly"):
            conn.close()
            return ResponseHelper.error("会员类型须为 normal / free / monthly / yearly", 400)
        updates.append("member_type = ?")
        params.append(body.member_type)
    if body.member_start_at is not None:
        updates.append("member_start_at = ?")
        params.append(body.member_start_at.strip() if body.member_start_at else None)
    if body.member_end_at is not None:
        updates.append("member_end_at = ?")
        params.append(body.member_end_at.strip() if body.member_end_at else None)
    elif body.member_type == "free":
        # 免费用户：若未传 member_end_at，按 member_start_at + 3 个月（约90天）计算
        cursor.execute(
            "SELECT member_start_at, member_end_at FROM users WHERE id = ?", (user_id,)
        )
        row = cursor.fetchone()
        start_at = body.member_start_at
        if start_at and isinstance(row, (tuple, list)) and row and row[0]:
            start_at = row[0]
        if isinstance(start_at, str):
            start_at = start_at.strip() or None
        if not start_at and row and row[0]:
            start_at = row[0]
        if not start_at:
            start_at = format_datetime()
        try:
            if "T" in str(start_at):
                dt = datetime.fromisoformat(str(start_at).replace("Z", "+00:00")[:19])
            else:
                dt = datetime.strptime(str(start_at)[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            dt = datetime.now()
        end_dt = dt + timedelta(days=90)
        end_at = end_dt.strftime("%Y-%m-%d %H:%M:%S")
        updates.append("member_end_at = ?")
        params.append(end_at)
        if "member_start_at" not in [u.split("=")[0].strip() for u in updates]:
            updates.append("member_start_at = ?")
            params.append(dt.strftime("%Y-%m-%d %H:%M:%S"))
    if body.recognition_used is not None:
        updates.append("recognition_used = ?")
        params.append(body.recognition_used)

    if not updates:
        conn.close()
        return ResponseHelper.success(None, "无变更")

    updates.append("updated_at = ?")
    params.append(format_datetime())
    params.append(user_id)
    cursor.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
    conn.commit()
    conn.close()
    return ResponseHelper.success(None, "更新成功")


@router.delete("/admin/user/{user_id}")
async def admin_delete_user(user_id: str):
    """删除用户（main.db 及用户目录、用户库）。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE id = ?", (user_id,))
    if not cursor.fetchone():
        conn.close()
        return ResponseHelper.error("用户不存在", 404)

    user_db_path = config.get_user_db_path(user_id)
    if os.path.exists(user_db_path):
        try:
            os.remove(user_db_path)
        except Exception:
            pass
    user_dir = config.get_user_dir(user_id)
    if os.path.exists(user_dir):
        try:
            shutil.rmtree(user_dir)
        except Exception:
            pass

    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return ResponseHelper.success({"deleted": True}, "删除成功")


@router.get("/admin/stats")
async def admin_get_stats():
    """管理端统计：用户总数等。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM users")
    total_users = cursor.fetchone()[0]
    cursor.execute(
        "SELECT COUNT(*) FROM users WHERE COALESCE(member_type,'normal') IN ('monthly','yearly')"
    )
    member_count = cursor.fetchone()[0]
    conn.close()
    return ResponseHelper.success(
        {"total_users": total_users, "member_count": member_count},
        "获取成功",
    )


# ---------- 发票管理（从 main.db invoice_sync 读，详情从用户库读） ----------

@router.get("/admin/invoices")
async def admin_get_invoices(page: int = 1, limit: int = 10, keyword: Optional[str] = None):
    """管理端发票列表，分页、关键词搜索（发票号/购买方/销售方/文件名）。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    params: List[Any] = []
    where = ""
    if keyword and keyword.strip():
        k = f"%{keyword.strip()}%"
        where = "WHERE (invoice_number LIKE ? OR buyer LIKE ? OR seller LIKE ? OR filename LIKE ?)"
        params = [k, k, k, k]
    offset = (page - 1) * limit
    cursor.execute(
        f"""SELECT s.id, s.user_id, s.invoice_id, s.batch_id, s.invoice_number, s.invoice_date,
                   s.invoice_amount, s.total_with_tax, s.buyer, s.seller, s.recognition_status,
                   s.processing_time, s.upload_time, s.file_type, s.file_size, s.filename,
                   COALESCE(u.phone, u.username, u.email, '') AS user_name
            FROM invoice_sync AS s
            LEFT JOIN users AS u ON s.user_id = u.id
            {where}
            ORDER BY s.upload_time DESC
            LIMIT ? OFFSET ?""",
        (*params, limit, offset),
    )
    rows = cursor.fetchall()
    cursor.execute(f"SELECT COUNT(*) FROM invoice_sync {where}", params)
    total = cursor.fetchone()[0]
    conn.close()
    invoices = [dict(r) for r in rows]
    return ResponseHelper.success(
        {"invoices": invoices, "total": total, "page": page, "limit": limit},
        "获取成功",
    )


def _get_user_db_connection(user_id: str):
    import sqlite3
    path = config.get_user_db_path(user_id)
    if not os.path.exists(path):
        return None
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn


@router.get("/invoice/detail/{invoice_id}")
async def admin_get_invoice_detail(invoice_id: str):
    """管理端发票详情。invoice_id 可为 sync 主键（user_id_invoice_id）或直接 invoice_id，需从 sync 表反查 user_id。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    sync_row = None
    # 优先按 sync 主键查（fp_admin 列表里传的就是 invoice_sync.id）
    cursor.execute(
        "SELECT * FROM invoice_sync WHERE id = ? LIMIT 1",
        (invoice_id,),
    )
    sync_row = cursor.fetchone()

    if sync_row:
        user_id, inv_id = sync_row["user_id"], sync_row["invoice_id"]
    elif "_" in invoice_id:
        # 兼容旧参数：user_id_invoice_id
        parts = invoice_id.split("_", 1)
        user_id, inv_id = parts[0], parts[1]
        cursor.execute(
            "SELECT * FROM invoice_sync WHERE user_id = ? AND invoice_id = ? LIMIT 1",
            (user_id, inv_id),
        )
        sync_row = cursor.fetchone()
    else:
        cursor.execute(
            "SELECT * FROM invoice_sync WHERE invoice_id = ? LIMIT 1",
            (invoice_id,),
        )
        sync_row = cursor.fetchone()
        if not sync_row:
            conn.close()
            return ResponseHelper.error("发票不存在", 404)
        user_id, inv_id = sync_row["user_id"], sync_row["invoice_id"]

    conn.close()

    uconn = _get_user_db_connection(user_id)
    data = None
    if uconn:
        uc = uconn.cursor()
        uc.execute("SELECT * FROM invoice_details WHERE id = ?", (inv_id,))
        inv = uc.fetchone()
        uconn.close()
        if inv:
            data = dict(inv)

    # 若用户库缺失或找不到详情，则退回 invoice_sync 的行
    if not data:
        data = dict(sync_row) if sync_row else {}

    base = config.API_BASE_URL
    path = data.get("original_file_path") or data.get("processed_file_path") or ""
    file_url = None
    if path:
        file_url = base + (path if path.startswith("/") else "/" + path)
    else:
        # fallback：拼接用户上传目录 + saved_filename
        saved = data.get("saved_filename") or (sync_row["saved_filename"] if sync_row else None)
        uid = data.get("user_id") or (sync_row["user_id"] if sync_row else None)
        if saved and uid:
            file_url = f"{base}/files/{uid}/uploads/{saved}"
    data["file_url"] = file_url
    return ResponseHelper.success(data, "获取成功")


@router.get("/admin/invoice-stats")
async def admin_get_invoice_stats():
    """管理端发票统计：总数、已识别数、总金额。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM invoice_sync")
    total_invoices = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM invoice_sync WHERE recognition_status = 1")
    recognized_invoices = cursor.fetchone()[0]
    cursor.execute(
        "SELECT COALESCE(SUM(CAST(COALESCE(total_with_tax, invoice_amount) AS REAL)), 0) FROM invoice_sync"
    )
    row = cursor.fetchone()
    total_amount = float(row[0] or 0) if row else 0.0
    conn.close()
    return ResponseHelper.success(
        {
            "total_invoices": total_invoices,
            "recognized_invoices": recognized_invoices,
            "total_amount": total_amount,
        },
        "获取成功",
    )


@router.get("/admin/invoices/export")
async def admin_export_invoices(keyword: Optional[str] = None):
    """导出发票列表为 Excel。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    params = []
    where = ""
    if keyword and keyword.strip():
        k = f"%{keyword.strip()}%"
        where = "WHERE (invoice_number LIKE ? OR buyer LIKE ? OR seller LIKE ? OR filename LIKE ?)"
        params = [k, k, k, k]
    cursor.execute(
        f"""SELECT invoice_number, invoice_date, invoice_amount, total_with_tax, buyer, seller,
                   recognition_status, upload_time, file_type
            FROM invoice_sync {where}
            ORDER BY upload_time DESC""",
        params,
    )
    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "发票列表"
    headers = ["发票号码", "开票日期", "金额", "价税合计", "购买方", "销售方", "识别状态", "上传时间", "文件类型"]
    ws.append(headers)
    for r in rows:
        r = dict(r)
        status = {0: "待识别", 1: "已识别", 2: "失败"}.get(r.get("recognition_status"), "未知")
        ws.append([
            r.get("invoice_number") or "",
            r.get("invoice_date") or "",
            r.get("invoice_amount") or "",
            r.get("total_with_tax") or "",
            r.get("buyer") or "",
            r.get("seller") or "",
            status,
            r.get("upload_time") or "",
            r.get("file_type") or "",
        ])
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="invoices_export.xlsx"'},
    )


@router.delete("/invoice/{invoice_id}")
async def admin_delete_invoice(invoice_id: str):
    """管理端删除发票：从 sync 表查 user_id，再删用户库中记录及 sync 记录。"""
    conn = get_main_db_connection()
    cursor = conn.cursor()
    if "_" in invoice_id:
        parts = invoice_id.split("_", 1)
        user_id, inv_id = parts[0], parts[1]
    else:
        cursor.execute("SELECT user_id, invoice_id FROM invoice_sync WHERE invoice_id = ? LIMIT 1", (invoice_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return ResponseHelper.error("发票不存在", 404)
        user_id, inv_id = row["user_id"], row["invoice_id"]
    cursor.execute("DELETE FROM invoice_sync WHERE id = ?", (f"{user_id}_{inv_id}",))
    conn.commit()
    conn.close()

    try:
        from fp_api_new.invoices import get_user_db_connection
    except Exception:
        from invoices import get_user_db_connection
    uconn = get_user_db_connection(user_id)
    uc = uconn.cursor()
    uc.execute("DELETE FROM invoice_details WHERE id = ?", (inv_id,))
    uconn.commit()
    uconn.close()
    return ResponseHelper.success({"deleted": True}, "删除成功")
