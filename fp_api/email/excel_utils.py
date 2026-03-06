# 将发票解析结果导出为 Excel（不包含原始内容列）
import os
from invoice_parser import INVOICE_COLUMNS

# 导出 Excel 时的列（去掉原始内容）
EXCEL_COLUMNS = [c for c in INVOICE_COLUMNS if c != "原始内容"]

def build_excel(rows, save_path):
    """
    rows: list of dict, 每个 dict 的 key 为 INVOICE_COLUMNS 中的列名。
    save_path: 保存的 xlsx 路径。
    导出时只写入 EXCEL_COLUMNS，不包含原始内容。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise RuntimeError("请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票明细"

    # 表头（仅 EXCEL_COLUMNS）
    for col, name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 数据行
    for r, row in enumerate(rows, 2):
        for c, key in enumerate(EXCEL_COLUMNS, 1):
            ws.cell(row=r, column=c, value=row.get(key, ""))

    # 列宽
    for col in range(1, len(EXCEL_COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    d = os.path.dirname(save_path)
    if d:
        os.makedirs(d, exist_ok=True)
    wb.save(save_path)
    return save_path
